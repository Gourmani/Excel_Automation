import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process(input_filename, output_filename):
    wb = xl.load_workbook(input_filename)  # Load the existing Excel file
    sheet = wb["Sheet1"]  # Select the sheet

    print(sheet["A1"].value)  # Print the value in cell A1 (just to check)

    m_row = sheet.max_row  # Get total number of rows

    for i in range(2, m_row + 1):  # Skip the first row (header)
        cell = sheet.cell(i, 3)  # Column 3 (Price)
        c_p = cell.value * 0.9  # Apply 10% discount
        sheet.cell(i, 4, c_p)  # Store the discounted price in column 4

    # Create a Bar Chart for discounted prices
    values = Reference(sheet, min_col=4, min_row=2, max_row=m_row, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "E2")  # Place the chart at E2

    wb.save(output_filename)  # Save the updated file with a new name
    print(f"Updated file saved as: {output_filename}")

# Example usage
process("transactions.xlsx", "updated_transactions.xlsx")
